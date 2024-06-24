from django.shortcuts import render

# Create your views here.

from rest_framework import viewsets, status
from rest_framework.decorators import api_view
from rest_framework.response import Response

from .models import *
from .serializers import *
import openpyxl
from http import HTTPStatus
from django.http import (HttpResponse, HttpResponseBadRequest, HttpResponseForbidden)
from django.http import JsonResponse
from django.core.exceptions import ObjectDoesNotExist
from django.db.models import Q
from decimal import Decimal

class BuildingViewSet(viewsets.ModelViewSet):
    queryset = Building.objects.all()
    serializer_class = BuildingSerializer


class ContractViewSet(viewsets.ModelViewSet):
    queryset = Contract.objects.all()
    serializer_class = ContractSerializer


class ContractBuildingConnectionViewSet(viewsets.ModelViewSet):
    queryset = ContractBuildingConnection.objects.all()
    serializer_class = ContractBuildingConnectionSerializer


class FixedAssetViewSet(viewsets.ModelViewSet):
    queryset = FixedAsset.objects.all()
    serializer_class = FixedAssetSerializer


class ServiceViewSet(viewsets.ModelViewSet):
    queryset = Service.objects.all()
    serializer_class = ServiceSerializer


class InvoiceForPaymentViewSet(viewsets.ModelViewSet):
    queryset = InvoiceForPayment.objects.all()
    serializer_class = InvoiceForPaymentSerializer


class DistributedInvoiceForPaymentViewSet(viewsets.ModelViewSet):
    queryset = DistributedInvoiceForPayment.objects.all()
    serializer_class = DistributedInvoiceForPaymentSerializer


class_lookup = {'building': Building, 'contract': Contract, 'contract_building_connection': ContractBuildingConnection, 'fixed_asset': FixedAsset,
                        'service': Service, 'invoice_for_payment': InvoiceForPayment, 'distributed_invoice_for_payment': DistributedInvoiceForPayment}


serializer_class_lookup = {'building': BuildingSerializer, 'contract': ContractSerializer, 'contract_building_connection': ContractBuildingConnectionSerializer,
                           'fixed_asset': FixedAssetSerializer, 'service': ServiceSerializer, 'invoice_for_payment': InvoiceForPaymentSerializer,
                           'distributed_invoice_for_payment': DistributedInvoiceForPaymentWithoutIdSerializer}


def deserialize_building(data):
    building = Building(building_id=data['building_id'], possession_beginning_date=data['possession_beginning_date'],
                        possession_ending_date=data['possession_ending_date'], measurement_ending_date=data['measurement_ending_date'],
                        measurement_beginning_date=data['measurement_beginning_date'], square=data['square'], measure_unit=data['measure_unit'])
    return building


def deserialize_contract(data):
    contract = Contract(contract_id=data['contract_id'], contract_beginning_date=data['contract_beginning_date'], contract_ending_date=data['contract_ending_date'])
    return contract


def deserialize_contract_building_connection(data):
    contract_building_connection = ContractBuildingConnection(contract_id=data['contract_id'], building_id=data['building_id'],
                                                              connection_beginning_date=data['connection_beginning_date'],
                                                              connection_ending_date=data['connection_ending_date'])
    return contract_building_connection


def deserialize_fixed_asset(data):
    fixed_asset = FixedAsset(fixed_asset_id=data['fixed_asset_id'], fixed_asset_class=data['fixed_asset_class'],
                             is_used_in_main_activity=data['is_used_in_main_activity'], is_used_in_rent=data['is_used_in_rent'],
                             square=data['square'], measure_unit=data['measure_unit'], building_id=data['building_id'],
                             connection_with_building_beginning_date=data['connection_with_building_beginning_date'],
                             connection_with_building_ending_date=data['connection_with_building_ending_date'],
                             place_in_service_date=data['place_in_service_date'], disposal_date=data['contract_id'])
    return fixed_asset


def deserialize_service(data):
    service = Service(service_id=data['service_id'], service_class=data['service_class'])
    return service


def deserialize_invoice_for_payment(data):
    invoice_for_payment = InvoiceForPayment(company=data['company'], year=data['year'], invoice_number=data['invoice_number'],
                                            invoice_position=data['invoice_position'], service_id=data['service_id'], contract_id=data['contract_id'],
                                            invoice_reflection_in_the_accounting_system_date=data['invoice_reflection_in_the_accounting_system_date'],
                                            cost_excluding_VAT=Decimal(data['cost_excluding_VAT']))
    return invoice_for_payment


def deserialize_distributed_invoice_for_payment(data):
    distributed_invoice_for_payment = DistributedInvoiceForPayment(company=data['company'], year=data['year'], invoice_number=data['invoice_number'],
                                                                   invoice_position=data['invoice_position'], distribution_position_number=data['distribution_position_number'],
                                                                   reflection_in_the_accounting_system_date=data['reflection_in_the_accounting_system_date'],
                                                                   contract_id=data['contract_id'], service_id=data['service_id'], service_class=data['service_class'],
                                                                   building_id=data['building_id'], fixed_asset_class=data['fixed_asset_class'], fixed_asset_id=data['fixed_asset_id'],
                                                                   is_used_in_main_activity=data['is_used_in_main_activity'], is_used_in_rent=data['is_used_in_rent'],
                                                                   square=data['square'], distribution_sum=data['distribution_sum'], general_ledger_account=data['general_ledger_account'])
    return distributed_invoice_for_payment


def upload_building_file(worksheet):
    buildings = []
    try:
        for row in worksheet.iter_rows(min_row=2):
                building = Building(building_id=row[0].value, possession_beginning_date=row[1].value, possession_ending_date=row[2].value, measurement_ending_date=row[3].value,
                                    measurement_beginning_date=row[4].value, square=row[5].value, measure_unit=row[6].value)
                buildings.append(building)
        Building.objects.bulk_create(buildings)
    except Exception as err:
        return Response(status=status.HTTP_422_UNPROCESSABLE_ENTITY)
    return HttpResponse(status=200)


def upload_contract_file(worksheet):
    contracts = []
    try:
        for row in worksheet.iter_rows(min_row=2):
            contract = Contract(contract_id=row[0].value, contract_beginning_date=row[1].value, contract_ending_date=row[2].value)
            contracts.append(contract)
        Contract.objects.bulk_create(contracts)
    except Exception as err:
        return Response(status=status.HTTP_422_UNPROCESSABLE_ENTITY)
    return HttpResponse(status=200)


def upload_contract_building_connection_file(worksheet):
    contract_building_connections = []
    try:
        for row in worksheet.iter_rows(min_row=2):
            contract_building_connection = ContractBuildingConnection(contract_id=row[0].value, building_id=row[1].value, connection_beginning_date=row[2].value,
                                                                      connection_ending_date=row[3].value)
            contract_building_connections.append(contract_building_connection)
        ContractBuildingConnection.objects.bulk_create(contract_building_connections)
    except Exception as err:
        return Response(status=status.HTTP_422_UNPROCESSABLE_ENTITY)
    return HttpResponse(status=200)


def upload_fixed_asset_file(worksheet):
    fixed_assets = []
    try:
        for row in worksheet.iter_rows(min_row=2):
            fixed_asset = FixedAsset(fixed_asset_id=row[0].value, fixed_asset_class=row[1].value, is_used_in_main_activity=(row[2].value == 'X'),
                                     is_used_in_rent=(row[3].value == 'X'), square=row[4].value, measure_unit=row[5].value, building_id=row[6].value,
                                     connection_with_building_beginning_date=row[7].value, connection_with_building_ending_date=row[8].value, place_in_service_date=row[9].value,
                                     disposal_date=row[10].value)
            fixed_assets.append(fixed_asset)
        FixedAsset.objects.bulk_create(fixed_assets)
    except Exception as err:
        return Response(status=status.HTTP_422_UNPROCESSABLE_ENTITY)
    return HttpResponse(status=200)


def upload_service_file(worksheet):
    services = []
    try:
        for row in worksheet.iter_rows(min_row=2):
            service = Service(service_id=row[0].value, service_class=row[1].value)
            services.append(service)
        Service.objects.bulk_create(services)
    except Exception as err:
        return Response(status=status.HTTP_422_UNPROCESSABLE_ENTITY)
    return HttpResponse(status=200)


def upload_invoice_for_payment_file(worksheet):
    invoice_for_payments = []
    try:
        for row in worksheet.iter_rows(min_row=2):
            invoice_for_payment = InvoiceForPayment(company=row[0].value, year=row[1].value, invoice_number=row[2].value, invoice_position=row[3].value, service_id=row[4].value,
                                                    contract_id=row[5].value, invoice_reflection_in_the_accounting_system_date=row[6].value, cost_excluding_VAT=row[7].value)
            invoice_for_payments.append(invoice_for_payment)
        InvoiceForPayment.objects.bulk_create(invoice_for_payments)
    except Exception as err:
        return Response(status=status.HTTP_422_UNPROCESSABLE_ENTITY)
    return HttpResponse(status=200)


def upload_distributed_invoice_for_payment_file(worksheet):
    distributed_invoice_for_payments = []
    try:
        for row in worksheet.iter_rows(min_row=2):
            distributed_invoice_for_payment = DistributedInvoiceForPayment(company=row[0].value, year=row[1].value, invoice_number=row[2].value, invoice_position=row[3].value,
                                                                           distribution_position_number=row[4].value, reflection_in_the_accounting_system_date=row[5].value,
                                                                           contract_id=row[6].value, service_id=row[7].value, service_class=row[8].value, building_id=row[9].value,
                                                                           fixed_asset_class=row[10].value, fixed_asset_id=row[11].value, is_used_in_main_activity=(row[12].value == 'X'),
                                                                           is_used_in_rent=(row[13].value == 'X'), square=row[14].value, distribution_sum=row[15].value,
                                                                           general_ledger_account=row[16].value)
            distributed_invoice_for_payments.append(distributed_invoice_for_payment)
        DistributedInvoiceForPayment.objects.bulk_create(distributed_invoice_for_payments)
    except Exception as err:
        return Response(status=status.HTTP_422_UNPROCESSABLE_ENTITY)
    return HttpResponse(status=200)


@api_view(['POST'])
def upload_file(request, table_name):
    excel_file = request.FILES["excel_file"]
    wb = openpyxl.load_workbook(excel_file)
    worksheet = wb["Sheet1"]
    response = globals()['upload_' + table_name + '_file'](worksheet)
    return response


@api_view(['POST'])
def upload_json(request, table_name):
    try:
        data = request.data
        objects_list = []
        for i in range(len(data)):
            objects_list.append(globals()['deserialize_' + table_name](data[i]))
        class_lookup[table_name].objects.bulk_create(objects_list)
        if (table_name == 'invoice_for_payment'):
            return start_distribution(objects_list)
    except Exception as err:
        return Response(status=status.HTTP_422_UNPROCESSABLE_ENTITY)
    return HttpResponse(status=201)


def start_distribution(invoice_for_payments):
    try:
        distributed_invoice_for_payments = []
        filter_services = Q()
        for invoice_for_payment in invoice_for_payments:
            filter_services = filter_services | Q(service_id=invoice_for_payment.service_id)
        services = Service.objects.filter(filter_services)

        filter_contract_building_connections = Q()
        for invoice_for_payment in invoice_for_payments:
            filter_contract_building_connections = filter_contract_building_connections | Q(contract_id=invoice_for_payment.contract_id)
        contract_building_connections = ContractBuildingConnection.objects.filter(filter_contract_building_connections)

        filter_fixed_assets = Q()
        for contract_building_connection in contract_building_connections:
            filter_fixed_assets = filter_fixed_assets | Q(building_id=contract_building_connection.building_id)
        fixed_assets = FixedAsset.objects.filter(filter_fixed_assets)

        for i in range(len(invoice_for_payments)):
            count = 0
            square_sum = 0
            invoice_for_payment = invoice_for_payments[i]
            company = invoice_for_payment.company
            year = invoice_for_payment.year
            invoice_number = invoice_for_payment.invoice_number
            invoice_position = invoice_for_payment.invoice_position
            service_id = invoice_for_payment.service_id
            service_class = list(filter(lambda service: service.service_id == service_id, services))[0].service_class
            reflection_in_the_accounting_system_date = invoice_for_payment.invoice_reflection_in_the_accounting_system_date
            cost_excluding_VAT = invoice_for_payment.cost_excluding_VAT
            contract_id = invoice_for_payment.contract_id
            temp_contract_building_connections = list(filter(lambda contract_building_connection: contract_building_connection.contract_id == contract_id, contract_building_connections))
            for j in range(len(temp_contract_building_connections)):
                building_id = temp_contract_building_connections[j].building_id
                temp_fixed_assets = list(filter(lambda fixed_asset: fixed_asset.building_id == building_id, fixed_assets))
                for k in range(len(temp_fixed_assets)):
                    fixed_asset_id = temp_fixed_assets[k].fixed_asset_id
                    fixed_asset_class = temp_fixed_assets[k].fixed_asset_class
                    is_used_in_main_activity = temp_fixed_assets[k].is_used_in_main_activity
                    is_used_in_rent = temp_fixed_assets[k].is_used_in_rent
                    square = temp_fixed_assets[k].square
                    distributed_invoice_for_payment = DistributedInvoiceForPayment(company=company, year=year, invoice_number=invoice_number, invoice_position=invoice_position,
                                                                                   distribution_position_number=count + 1,
                                                                                   reflection_in_the_accounting_system_date=reflection_in_the_accounting_system_date,
                                                                                   contract_id=contract_id, service_id=service_id, service_class=service_class, building_id=building_id,
                                                                                   fixed_asset_class=fixed_asset_class, fixed_asset_id=fixed_asset_id,
                                                                                   is_used_in_main_activity=is_used_in_main_activity, is_used_in_rent=is_used_in_rent,
                                                                                   square=square, distribution_sum=0)
                    distributed_invoice_for_payments.append(distributed_invoice_for_payment)
                    count += 1
                    square_sum += square
            for j in range(count):
                distributed_invoice_for_payments[len(distributed_invoice_for_payments) - count + j].distribution_sum = cost_excluding_VAT * distributed_invoice_for_payments[len(distributed_invoice_for_payments) - count + j].square / square_sum
        serializer = DistributedInvoiceForPaymentWithoutIdSerializer(distributed_invoice_for_payments, many=True)
    except Exception as err:
        return Response(status=status.HTTP_422_UNPROCESSABLE_ENTITY)
    return Response(serializer.data, status=status.HTTP_200_OK)


def index(request):
    selected_object = str(request.POST.get('selected_object'))
    if selected_object == "None":
        selected_object = 'building'
    if 'delete' in request.POST:
        class_lookup[selected_object].objects.all().delete()
    return render(request, 'index.html', {"object": selected_object})